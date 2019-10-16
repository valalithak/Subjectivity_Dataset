#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import requests
import datetime
from time import sleep
import os
import commands
import re

wb = load_workbook("subjectivity_annotation.xlsx", data_only=True)
sh = wb["Sheet1"]
i = 2
while sh["b"+str(i)].value != None:
	print(i)
	article = sh["b" + str(i)].value
	print(article)
	f = open('temp.txt' , 'w')
	f.write(article.encode('utf8'))
	f.write("\n")
	f.close()

	subj_val = commands.getoutput("python -m subjectivity.classify < temp.txt")

	# cell = sh.cell(row = i, column = 3)
	# cell.value = c
	print(subj_val)
	print("\n")

	i +=1 

# wb.save("subjectivity_annotation.xlsx")
